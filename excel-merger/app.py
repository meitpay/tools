import os
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
from copy import copy
from io import BytesIO

app = Flask(__name__, static_folder="static", template_folder="templates")

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        files = request.files.getlist("files")
        include_review = bool(request.form.get("include_review"))

        if not files:
            return "No files uploaded.", 400

        merged_wb = Workbook()
        merged_wb.remove(merged_wb.active)

        for f in files:
            filename = secure_filename(f.filename)
            stream = BytesIO(f.read())
            wb = load_workbook(stream, data_only=False)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                new_ws = merged_wb.create_sheet(f"{filename[:-5]}_{sheet_name}")

                max_row = ws.max_row
                max_col = ws.max_column

                for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                    for cell in row:
                        col_offset = 1 if include_review else 0
                        new_cell = new_ws.cell(
                            row=cell.row,
                            column=cell.column + col_offset,
                            value=cell.value
                        )

                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.fill = copy(cell.fill)
                            new_cell.border = copy(cell.border)
                            new_cell.number_format = cell.number_format
                            new_cell.alignment = copy(cell.alignment)

                        if cell.comment:
                            new_cell.comment = copy(cell.comment)

                if include_review:
                    for row_idx in range(1, max_row + 1):
                        new_ws.cell(
                            row=row_idx,
                            column=1,
                            value="Review" if row_idx == 1 else ""
                        )

                filter_end_col = get_column_letter(max_col + 1 if include_review else max_col)
                new_ws.auto_filter.ref = f"A1:{filter_end_col}1"

                total_cols = max_col + 1 if include_review else max_col
                for col in range(1, total_cols + 1):
                    max_len = 0
                    col_letter = get_column_letter(col)
                    for cell in new_ws[col_letter]:
                        try:
                            if cell.value:
                                max_len = max(max_len, len(str(cell.value)))
                        except:
                            pass
                    new_ws.column_dimensions[col_letter].width = max_len + 2

        output = BytesIO()
        merged_wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="merged.xlsx")

    return render_template("index.html")


if __name__ == "__main__":
    app.run(debug=True)
